"""
Excel Report Automation
--------------------------------------------------
Toma un archivo CSV de datos crudos (ej. ventas) y genera
automáticamente un reporte en Excel con:
  - Formato profesional
  - Totales y resumen
  - Gráfica de barras

Autor: [Tu nombre]
"""

import pandas as pd
from pathlib import Path
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# --------------------------------------------------------------
# CONFIGURACIÓN
# --------------------------------------------------------------
INPUT_FILE = Path(__file__).parent / "datos_ejemplo.csv"
OUTPUT_FILE = Path(__file__).parent / "reporte_generado.xlsx"

# Columna sobre la que se agrupa (ej. "producto", "vendedor", "region")
GROUP_BY_COLUMN = "producto"
# Columna numérica a sumar (ej. "ventas", "monto")
VALUE_COLUMN = "ventas"


def load_data(path: Path) -> pd.DataFrame:
    """Carga los datos crudos desde un CSV."""
    df = pd.read_csv(path)
    return df


def build_summary(df: pd.DataFrame) -> pd.DataFrame:
    """Agrupa y suma los datos para crear el resumen."""
    summary = (
        df.groupby(GROUP_BY_COLUMN)[VALUE_COLUMN]
        .sum()
        .reset_index()
        .sort_values(by=VALUE_COLUMN, ascending=False)
    )
    return summary


def style_header(ws, row: int, n_cols: int):
    """Aplica formato de encabezado (color, negrita) a una fila."""
    header_fill = PatternFill(start_color="2E5395", end_color="2E5395", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    for col in range(1, n_cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")


def autofit_columns(ws):
    """Ajusta el ancho de columnas según el contenido."""
    for col_cells in ws.columns:
        length = max(len(str(cell.value)) if cell.value is not None else 0 for cell in col_cells)
        col_letter = get_column_letter(col_cells[0].column)
        ws.column_dimensions[col_letter].width = length + 4


def generate_report(df: pd.DataFrame, summary: pd.DataFrame, output_path: Path):
    """Genera el archivo Excel final con datos, resumen y gráfica."""
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="Datos", index=False)
        summary.to_excel(writer, sheet_name="Resumen", index=False)

        wb = writer.book

        # --- Formato hoja Datos ---
        ws_data = writer.sheets["Datos"]
        style_header(ws_data, 1, len(df.columns))
        autofit_columns(ws_data)

        # --- Formato hoja Resumen ---
        ws_summary = writer.sheets["Resumen"]
        style_header(ws_summary, 1, len(summary.columns))
        autofit_columns(ws_summary)

        # --- Gráfica de barras en hoja Resumen ---
        chart = BarChart()
        chart.title = f"Total de {VALUE_COLUMN} por {GROUP_BY_COLUMN}"
        chart.y_axis.title = VALUE_COLUMN
        chart.x_axis.title = GROUP_BY_COLUMN

        n_rows = len(summary) + 1
        data_ref = Reference(ws_summary, min_col=2, min_row=1, max_row=n_rows)
        cats_ref = Reference(ws_summary, min_col=1, min_row=2, max_row=n_rows)
        chart.add_data(data_ref, titles_from_data=True)
        chart.set_categories(cats_ref)
        ws_summary.add_chart(chart, "E2")


def run():
    print(f"Leyendo datos desde: {INPUT_FILE}")
    df = load_data(INPUT_FILE)

    print("Generando resumen...")
    summary = build_summary(df)

    print(f"Creando reporte en: {OUTPUT_FILE}")
    generate_report(df, summary, OUTPUT_FILE)

    print("\n✅ Reporte generado con éxito.")
    print(f"Total general de {VALUE_COLUMN}: {summary[VALUE_COLUMN].sum()}")


if __name__ == "__main__":
    run()
